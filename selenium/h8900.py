# pip install selenium
# pip install openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook

options = webdriver.ChromeOptions()
options.add_argument("--disable-proxy-certificate-handler")
options.add_argument("--allow-running-insecure-content")
options.add_argument("--ignore-certificate-errors")

# 打开要处理的页面
# driver = webdriver.Chrome()
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(10)
driver.maximize_window()

driver.get("https://111.222.333.444:5555/WPMS/")
driver.find_element(By.ID, "user-name").send_keys("user")
driver.find_element(By.ID, "password").send_keys("pass")
driver.find_element(By.ID, "login-submit").click()
driver.find_element(By.XPATH, '//*[@value="确定"]').click()
driver.find_element(By.XPATH, '//*[@id="home-page"]/a').click()
driver.find_element(By.XPATH, '//*[@name="人脸应用"]').click()
time.sleep(10)
driver.switch_to.frame('iframe-FACE')
driver.find_element(By.XPATH, '//*[@placeholder="模糊查询"]').send_keys("东湖花园白名单")
driver.find_element(By.XPATH, '//*[@placeholder="模糊查询"]/../div/button[1]').click()
driver.find_element(By.XPATH, '//button[contains(text(), "人员信息")]').click()

# 加载工作簿
wb = load_workbook('work.xlsx')
# 获取sheet页
ws = wb['person_chengxiang']
for i in range(2, 2644):  # 2是Excel表格的开始行，2644是结束行
    name = ws.cell(row=i, column=9).value  # 9是要处理的人员名字所在列
    driver.find_element(By.XPATH, '//*[@placeholder="模糊查询"]').clear()
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@placeholder="模糊查询"]').send_keys(name)
    driver.find_element(By.XPATH, '//*[@placeholder="模糊查询"]/../div/button[1]').click()
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@title="删除"]').click()
    time.sleep(1)
    driver.switch_to.default_content()
    driver.find_element(By.XPATH, '//*[@value="确定删除"]').click()
    driver.switch_to.frame('iframe-FACE')
    # driver.find_element(By.XPATH, '//*[@placeholder="模糊查询"]').clear()
    # time.sleep(1)
    # driver.find_element(By.XPATH, '//*[@placeholder="模糊查询"]').send_keys(name)
    # driver.find_element(By.XPATH, '//*[@placeholder="模糊查询"]/../div/button[1]').click()
    # time.sleep(1)
    print('已处理：' + name)
    ws.cell(row=i, column=12).value ='1'
    wb.save('work.xlsx')
wb.close()

driver.close()